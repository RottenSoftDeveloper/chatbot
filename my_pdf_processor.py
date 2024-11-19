import os
from dotenv import load_dotenv
from PyPDF2 import PdfReader
import docx
import openpyxl
from langchain.text_splitter import CharacterTextSplitter
from langchain.embeddings.openai import OpenAIEmbeddings
from langchain.vectorstores import FAISS
from langchain.chains.question_answering import load_qa_chain
from langchain.llms import OpenAI
from langchain.callbacks import get_openai_callback

load_dotenv(r'C:\Users\Fabian\python_files\.env')

############ TEXT LOADERS ############
def read_pdf(file_path):
    with open(file_path, "rb") as file:
        pdf_reader = PdfReader(file)
        text = ""
        for page_num in range(len(pdf_reader.pages)):
            text += pdf_reader.pages[page_num].extract_text()
    return text

def read_word(file_path):
    doc = docx.Document(file_path)
    text = ""
    for paragraph in doc.paragraphs:
        text += paragraph.text + "\n"
    return text

def read_txt(file_path):
    with open(file_path, "r") as file:
        text = file.read()
    return text

def read_excel(file_path):
    workbook = openpyxl.load_workbook(file_path, data_only=True)
    text = ""
    for sheet in workbook.sheetnames:
        worksheet = workbook[sheet]
        for row in worksheet.iter_rows(values_only=True):
            text += " ".join(str(cell) for cell in row if cell is not None) + "\n"
    return text

def read_documents_from_directory(directory):
    combined_text = ""
    for filename in os.listdir(directory):
        file_path = os.path.join(directory, filename)
        if filename.endswith(".pdf"):
            combined_text += read_pdf(file_path)
        elif filename.endswith(".docx"):
            combined_text += read_word(file_path)
        elif filename.endswith(".txt"):
            combined_text += read_txt(file_path)
        elif filename.endswith(".xlsx"):
            combined_text += read_excel(file_path)
    return combined_text

###############################################

train_directory = 'train_files/'
text = read_documents_from_directory(train_directory)

# Split into chunks
char_text_splitter = CharacterTextSplitter(separator="\n", chunk_size=1000, 
                                      chunk_overlap=200, length_function=len)

text_chunks = char_text_splitter.split_text(text)

# Create embeddings
embeddings = OpenAIEmbeddings()
docsearch = FAISS.from_texts(text_chunks, embeddings)

llm = OpenAI()
chain = load_qa_chain(llm, chain_type="stuff")

##################################################

def clear_console():
    os.system('cls' if os.name == 'nt' else 'clear')
while True:
    query = input("Enter your question: ")
    if query.lower() == 'exit':
        break

    docs = docsearch.similarity_search(query)

    with get_openai_callback() as cb:
        response = chain.run(input_documents=docs, question=query)

    # Check for common uncertainty phrases
    if any(phrase in response.lower() for phrase in ["i don't know", "idk", "there is no mention", "unable to answer"]):
        response = "Please reach out to security validation at (321) 521-4387 option 5 for helpdesk."
    
    print("\nQuery:", query)
    print("Response:", response)
    print(cb)

    input("\nPress Enter to continue...")
    clear_console()


print("Exiting...")
